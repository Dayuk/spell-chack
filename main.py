from PyQt5.QtWidgets import QDesktopWidget
from PyQt5.QtWidgets import QLabel
import sys
from PyQt5.QtCore import QTimer
from PyQt5.QtWidgets import QProgressBar
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import *
from pynput import keyboard
from PyQt5.QtGui import *
from PyQt5.QtGui import QIcon
from PyQt5 import uic
import openpyxl

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.a = ['3000', '1800', '3600', '1800', '2400', '2100', '1800', '900', '2100']
        self.T = ['301', '181', '361', '181', '241', '211', '181', '91', '211']
        self.wb = openpyxl.Workbook()
        self.file = openpyxl.load_workbook("Key Data.xlsx")
        self.sheet = self.file.active
        self.key1 = self.sheet['A1'].value
        self.key3 = self.sheet['A2'].value
        self.key5 = self.sheet['A3'].value
        self.key7 = self.sheet['A4'].value
        self.key9 = self.sheet['A5'].value
        self.key2 = self.sheet['B1'].value
        self.key4 = self.sheet['B2'].value
        self.key6 = self.sheet['B3'].value
        self.key8 = self.sheet['B4'].value
        self.key10 = self.sheet['B5'].value
        self.sheet = self.wb.active

        # self.line1 = QLabel('TOP', self)
        # self.line1.move(25, 4)
        #
        # self.line1 = QLabel('JG', self)
        # self.line1.move(105, 4)
        #
        # self.line1 = QLabel('MID', self)
        # self.line1.move(175, 4)
        #
        # self.line1 = QLabel('AD', self)
        # self.line1.move(255, 4)
        #
        # self.line1 = QLabel('SUP', self)
        # self.line1.move(325, 4)

        self.pbar = QProgressBar(self)
        self.pbar.setAlignment(Qt.AlignCenter)
        self.pbar.setGeometry(5, 5, 68, 52)
        self.Dbtn1 = QPushButton('D spell', self)
        self.Dbtn1.resize(70, 55)
        self.Dbtn1.move(4, 4)
        self.Dbtn1.setCheckable(True)
        self.Dbtn1.clicked.connect(self.Dbtnn)
        self.Dbtn1.clicked.connect(self.timer1)

        self.pbar3 = QProgressBar(self)
        self.pbar3.setAlignment(Qt.AlignCenter)
        self.pbar3.setGeometry(80, 5, 68, 52)
        self.Dbtn2 = QPushButton('D spell', self)
        self.Dbtn2.resize(70, 55)
        self.Dbtn2.move(79, 4)
        self.Dbtn2.setCheckable(True)
        self.Dbtn2.clicked.connect(self.Dbtnn2)
        self.Dbtn2.clicked.connect(self.timer2)

        self.pbar4 = QProgressBar(self)
        self.pbar4.setAlignment(Qt.AlignCenter)
        self.pbar4.setGeometry(155, 5, 68, 52)
        self.Dbtn3 = QPushButton('D spell', self)
        self.Dbtn3.resize(70, 55)
        self.Dbtn3.move(154, 4)
        self.Dbtn3.setCheckable(True)
        self.Dbtn3.clicked.connect(self.Dbtnn3)
        self.Dbtn3.clicked.connect(self.timer3)

        self.pbar5 = QProgressBar(self)
        self.pbar5.setAlignment(Qt.AlignCenter)
        self.pbar5.setGeometry(230, 5, 68, 52)
        self.Dbtn4 = QPushButton('D spell', self)
        self.Dbtn4.resize(70, 55)
        self.Dbtn4.move(229, 4)
        self.Dbtn4.setCheckable(True)
        self.Dbtn4.clicked.connect(self.Dbtnn4)
        self.Dbtn4.clicked.connect(self.timer4)

        self.pbar6 = QProgressBar(self)
        self.pbar6.setAlignment(Qt.AlignCenter)
        self.pbar6.setGeometry(305, 5, 68, 52)
        self.Dbtn5 = QPushButton('D spell', self)
        self.Dbtn5.resize(70, 55)
        self.Dbtn5.move(304, 4)
        self.Dbtn5.setCheckable(True)
        self.Dbtn5.clicked.connect(self.Dbtnn5)
        self.Dbtn5.clicked.connect(self.timer5)

        self.pbar2 = QProgressBar(self)
        self.pbar2.setAlignment(Qt.AlignCenter)
        self.pbar2.setGeometry(5, 61, 68, 52)
        self.Fbtn = QPushButton('F spell', self)
        self.Fbtn.resize(70, 55)
        self.Fbtn.move(4, 60)
        self.Fbtn.setCheckable(True)
        self.Fbtn.clicked.connect(self.Fbtnn)
        self.Fbtn.clicked.connect(self.timer6)

        self.pbarr = QProgressBar(self)
        self.pbarr.setAlignment(Qt.AlignCenter)
        self.pbarr.setGeometry(80, 61, 68, 52)
        self.Fbtn2 = QPushButton('F spell', self)
        self.Fbtn2.resize(70, 55)
        self.Fbtn2.move(79, 60)
        self.Fbtn2.setCheckable(True)
        self.Fbtn2.clicked.connect(self.Fbtnn2)
        self.Fbtn2.clicked.connect(self.timer7)

        self.pbarr2 = QProgressBar(self)
        self.pbarr2.setAlignment(Qt.AlignCenter)
        self.pbarr2.setGeometry(155, 61, 68, 52)
        self.Fbtn3 = QPushButton('F spell', self)
        self.Fbtn3.resize(70, 55)
        self.Fbtn3.move(154, 60)
        self.Fbtn3.setCheckable(True)
        self.Fbtn3.clicked.connect(self.Fbtnn3)
        self.Fbtn3.clicked.connect(self.timer8)

        self.pbarr3 = QProgressBar(self)
        self.pbarr3.setAlignment(Qt.AlignCenter)
        self.pbarr3.setGeometry(230, 61, 68, 52)
        self.Fbtn4 = QPushButton('F spell', self)
        self.Fbtn4.resize(70, 55)
        self.Fbtn4.move(229, 60)
        self.Fbtn4.setCheckable(True)
        self.Fbtn4.clicked.connect(self.Fbtnn4)
        self.Fbtn4.clicked.connect(self.timer9)

        self.pbarr4 = QProgressBar(self)
        self.pbarr4.setAlignment(Qt.AlignCenter)
        self.pbarr4.setGeometry(305, 61, 68, 52)
        self.Fbtn5 = QPushButton('F spell', self)
        self.Fbtn5.resize(70, 55)
        self.Fbtn5.move(304, 60)
        self.Fbtn5.setCheckable(True)
        self.Fbtn5.clicked.connect(self.Fbtnn5)
        self.Fbtn5.clicked.connect(self.timer10)

        self.Reset = QPushButton('Reset', self)
        self.Reset.resize(68, 55)
        self.Reset.move(380, 128)
        self.Reset.setCheckable(True)
        self.Reset.clicked.connect(self.ResetBtn)

        self.dcb = QComboBox(self)
        self.dcb.addItem('D spell')
        self.dcb.addItem('점멸')
        self.dcb.addItem('점화')
        self.dcb.addItem('텔레포트')
        self.dcb.addItem('유체화')
        self.dcb.addItem('회복')
        self.dcb.addItem('탈진')
        self.dcb.addItem('방어막')
        self.dcb.addItem('강타')
        self.dcb.addItem('정화')
        self.dcb.resize(68, 22)
        self.dcb.move(5, 130)
        self.dcb.currentIndexChanged.connect(self.ddsp)
        self.dcb.activated[str].connect(self.doac1)

        self.fcb = QComboBox(self)
        self.fcb.addItem('F spell')
        self.fcb.addItem('점멸')
        self.fcb.addItem('점화')
        self.fcb.addItem('텔레포트')
        self.fcb.addItem('유체화')
        self.fcb.addItem('회복')
        self.fcb.addItem('탈진')
        self.fcb.addItem('방어막')
        self.fcb.addItem('강타')
        self.fcb.addItem('정화')
        self.fcb.resize(68, 22)
        self.fcb.move(5, 160)
        self.fcb.currentIndexChanged.connect(self.ffsp)
        self.fcb.activated[str].connect(self.foac1)

        self.dcb2 = QComboBox(self)
        self.dcb2.addItem('D spell')
        self.dcb2.addItem('점멸')
        self.dcb2.addItem('점화')
        self.dcb2.addItem('텔레포트')
        self.dcb2.addItem('유체화')
        self.dcb2.addItem('회복')
        self.dcb2.addItem('탈진')
        self.dcb2.addItem('방어막')
        self.dcb2.addItem('강타')
        self.dcb2.addItem('정화')
        self.dcb2.move(80, 130)
        self.dcb2.resize(68, 22)
        self.dcb2.currentIndexChanged.connect(self.ddsp2)
        self.dcb2.activated[str].connect(self.doac2)

        self.fcb2 = QComboBox(self)
        self.fcb2.addItem('F spell')
        self.fcb2.addItem('점멸')
        self.fcb2.addItem('점화')
        self.fcb2.addItem('텔레포트')
        self.fcb2.addItem('유체화')
        self.fcb2.addItem('회복')
        self.fcb2.addItem('탈진')
        self.fcb2.addItem('방어막')
        self.fcb2.addItem('강타')
        self.fcb2.addItem('정화')
        self.fcb2.move(80, 160)
        self.fcb2.resize(68, 22)
        self.fcb2.currentIndexChanged.connect(self.ffsp2)
        self.fcb2.activated[str].connect(self.foac2)

        self.dcb3 = QComboBox(self)
        self.dcb3.addItem('D spell')
        self.dcb3.addItem('점멸')
        self.dcb3.addItem('점화')
        self.dcb3.addItem('텔레포트')
        self.dcb3.addItem('유체화')
        self.dcb3.addItem('회복')
        self.dcb3.addItem('탈진')
        self.dcb3.addItem('방어막')
        self.dcb3.addItem('강타')
        self.dcb3.addItem('정화')
        self.dcb3.move(155, 130)
        self.dcb3.resize(68, 22)
        self.dcb3.currentIndexChanged.connect(self.ddsp3)
        self.dcb3.activated[str].connect(self.doac3)

        self.fcb3 = QComboBox(self)
        self.fcb3.addItem('F spell')
        self.fcb3.addItem('점멸')
        self.fcb3.addItem('점화')
        self.fcb3.addItem('텔레포트')
        self.fcb3.addItem('유체화')
        self.fcb3.addItem('회복')
        self.fcb3.addItem('탈진')
        self.fcb3.addItem('방어막')
        self.fcb3.addItem('강타')
        self.fcb3.addItem('정화')
        self.fcb3.move(155, 160)
        self.fcb3.resize(68, 22)
        self.fcb3.currentIndexChanged.connect(self.ffsp3)
        self.fcb3.activated[str].connect(self.foac3)

        self.dcb4 = QComboBox(self)
        self.dcb4.addItem('D spell')
        self.dcb4.addItem('점멸')
        self.dcb4.addItem('점화')
        self.dcb4.addItem('텔레포트')
        self.dcb4.addItem('유체화')
        self.dcb4.addItem('회복')
        self.dcb4.addItem('탈진')
        self.dcb4.addItem('방어막')
        self.dcb4.addItem('강타')
        self.dcb4.addItem('정화')
        self.dcb4.move(230, 130)
        self.dcb4.resize(68, 22)
        self.dcb4.currentIndexChanged.connect(self.ddsp4)
        self.dcb4.activated[str].connect(self.doac4)

        self.fcb4 = QComboBox(self)
        self.fcb4.addItem('F spell')
        self.fcb4.addItem('점멸')
        self.fcb4.addItem('점화')
        self.fcb4.addItem('텔레포트')
        self.fcb4.addItem('유체화')
        self.fcb4.addItem('회복')
        self.fcb4.addItem('탈진')
        self.fcb4.addItem('방어막')
        self.fcb4.addItem('강타')
        self.fcb4.addItem('정화')
        self.fcb4.move(230, 160)
        self.fcb4.resize(68, 22)
        self.fcb4.currentIndexChanged.connect(self.ffsp4)
        self.fcb4.activated[str].connect(self.foac4)

        self.dcb5 = QComboBox(self)
        self.dcb5.addItem('D spell')
        self.dcb5.addItem('점멸')
        self.dcb5.addItem('점화')
        self.dcb5.addItem('텔레포트')
        self.dcb5.addItem('유체화')
        self.dcb5.addItem('회복')
        self.dcb5.addItem('탈진')
        self.dcb5.addItem('방어막')
        self.dcb5.addItem('강타')
        self.dcb5.addItem('정화')
        self.dcb5.move(305, 130)
        self.dcb5.resize(68, 22)
        self.dcb5.currentIndexChanged.connect(self.ddsp5)
        self.dcb5.activated[str].connect(self.doac5)

        self.fcb5 = QComboBox(self)
        self.fcb5.addItem('F spell')
        self.fcb5.addItem('점멸')
        self.fcb5.addItem('점화')
        self.fcb5.addItem('텔레포트')
        self.fcb5.addItem('유체화')
        self.fcb5.addItem('회복')
        self.fcb5.addItem('탈진')
        self.fcb5.addItem('방어막')
        self.fcb5.addItem('강타')
        self.fcb5.addItem('정화')
        self.fcb5.move(305, 160)
        self.fcb5.resize(68, 22)
        self.fcb5.currentIndexChanged.connect(self.ffsp5)
        self.fcb5.activated[str].connect(self.foac5)

        label = QLabel('Made-By : LOLTOPIA', self)
        label.setAlignment(Qt.AlignVCenter)
        font1 = label.font()
        font1.setPointSize(15)
        layout = QVBoxLayout()
        layout.addWidget(label)
        label.move(5, 220)

        self.label1 = QLabel('000', self)
        self.label1.setAlignment(Qt.AlignVCenter)
        font1 = self.label1.font()
        font1.setPointSize(15)
        layout1 = QVBoxLayout()
        layout1.addWidget(self.label1)
        self.label1.move(30, 45)

        self.label2 = QLabel('000', self)
        self.label2.setAlignment(Qt.AlignVCenter)
        font2 = self.label2.font()
        font2.setPointSize(15)
        layout2 = QVBoxLayout()
        layout2.addWidget(self.label2)
        self.label2.move(105, 45)

        self.label3 = QLabel('000', self)
        self.label3.setAlignment(Qt.AlignVCenter)
        font3 = self.label3.font()
        font3.setPointSize(15)
        layout3 = QVBoxLayout()
        layout3.addWidget(self.label3)
        self.label3.move(180, 45)

        self.label4 = QLabel('000', self)
        self.label4.setAlignment(Qt.AlignVCenter)
        font4 = self.label4.font()
        font4.setPointSize(15)
        layout4 = QVBoxLayout()
        layout4.addWidget(self.label4)
        self.label4.move(255, 45)

        self.label5 = QLabel('000', self)
        self.label5.setAlignment(Qt.AlignVCenter)
        font5 = self.label5.font()
        font5.setPointSize(15)
        layout5 = QVBoxLayout()
        layout5.addWidget(self.label5)
        self.label5.move(330, 45)

        self.label6 = QLabel('000', self)
        self.label6.setAlignment(Qt.AlignVCenter)
        font6 = self.label6.font()
        font6.setPointSize(15)
        layout6 = QVBoxLayout()
        layout6.addWidget(self.label6)
        self.label6.move(30, 100)

        self.label7 = QLabel('000', self)
        self.label7.setAlignment(Qt.AlignVCenter)
        font7 = self.label7.font()
        font7.setPointSize(15)
        layout7 = QVBoxLayout()
        layout7.addWidget(self.label7)
        self.label7.move(105, 100)

        self.label8 = QLabel('000', self)
        self.label8.setAlignment(Qt.AlignVCenter)
        font8 = self.label8.font()
        font8.setPointSize(15)
        layout8 = QVBoxLayout()
        layout8.addWidget(self.label8)
        self.label8.move(180, 100)

        self.label9 = QLabel('000', self)
        self.label9.setAlignment(Qt.AlignVCenter)
        font9 = self.label9.font()
        font9.setPointSize(15)
        layout9 = QVBoxLayout()
        layout9.addWidget(self.label9)
        self.label9.move(255, 100)

        self.label10 = QLabel('000', self)
        self.label10.setAlignment(Qt.AlignVCenter)
        font10 = self.label10.font()
        font10.setPointSize(15)
        layout10 = QVBoxLayout()
        layout10.addWidget(self.label10)
        self.label10.move(330, 100)

        self.cb = QCheckBox('GMode', self)
        self.cb.move(380, 22)
        self.cb.stateChanged.connect(self.GameMode)

        self.cb2 = QCheckBox('KMode', self)
        self.cb2.move(380, 67)
        self.cb2.stateChanged.connect(self.keyEvent)

        self.slider = QSlider(Qt.Horizontal, self)
        self.slider.move(5, 190)
        self.slider.resize(130, 20)
        self.slider.setRange(0, 10)
        self.slider.setSingleStep(1)
        self.slider.NoTicks
        self.slider.valueChanged.connect(self.setValue)
        self.slider.setValue(0)

        self.Cbtn = QPushButton('Change\nInput Key', self)
        self.Cbtn.resize(68, 50)
        self.Cbtn.move(380, 185)
        self.Cbtn.setCheckable(True)
        self.Cbtn.clicked.connect(self.Change)

        self.setWindowTitle('스펠계산기')
        self.setWindowIcon(QIcon('img\\icon.png'))
        self.setWindowOpacity(1)
        self.setFixedSize(450, 238)
        self.center()
        self.show()

    def Change(self):
        self.Cbtn.toggle()
        self.dialog = QDialog()

        self.labelr1 = QLabel('D1 Key:', self.dialog)
        self.labelr1.move(10,10)

        self.keyy1 = QLineEdit(self.dialog)
        self.keyy1.resize(20,20)
        self.keyy1.move(54, 7)

        self.labelr2 = QLabel('D2 Key:', self.dialog)
        self.labelr2.move(90,10)

        self.keyy2 = QLineEdit(self.dialog)
        self.keyy2.resize(20,20)
        self.keyy2.move(134, 7)

        self.labelr3 = QLabel('D3 Key:', self.dialog)
        self.labelr3.move(170,10)

        self.keyy3 = QLineEdit(self.dialog)
        self.keyy3.resize(20,20)
        self.keyy3.move(214, 7)

        self.labelr4 = QLabel('D4 Key:', self.dialog)
        self.labelr4.move(250,10)

        self.keyy4 = QLineEdit(self.dialog)
        self.keyy4.resize(20,20)
        self.keyy4.move(294, 7)

        self.labelr5 = QLabel('D5 Key:', self.dialog)
        self.labelr5.move(330,10)

        self.keyy5 = QLineEdit(self.dialog)
        self.keyy5.resize(20,20)
        self.keyy5.move(374, 7)

        self.labelr6 = QLabel('F1 Key:', self.dialog)
        self.labelr6.move(10,70)
        self.keyy6 = QLineEdit(self.dialog)
        self.keyy6.resize(20,20)
        self.keyy6.move(54, 67)

        self.labelr7 = QLabel('F2 Key:', self.dialog)
        self.labelr7.move(90,70)
        self.keyy7 = QLineEdit(self.dialog)
        self.keyy7.resize(20,20)
        self.keyy7.move(134, 67)

        self.labelr8 = QLabel('F3 Key:', self.dialog)
        self.labelr8.move(170,70)
        self.keyy8 = QLineEdit(self.dialog)
        self.keyy8.resize(20,20)
        self.keyy8.move(214, 67)

        self.labelr9 = QLabel('F4 Key:', self.dialog)
        self.labelr9.move(250,70)
        self.keyy9 = QLineEdit(self.dialog)
        self.keyy9.resize(20,20)
        self.keyy9.move(294, 67)

        self.labelr10 = QLabel('F5 Key:', self.dialog)
        self.labelr10.move(330,70)
        self.keyy10 = QLineEdit(self.dialog)
        self.keyy10.resize(20,20)
        self.keyy10.move(374, 67)

        self.btnDialog1 = QPushButton("Save Key", self.dialog)
        self.btnDialog1.resize(70,25)
        self.btnDialog1.move(5, 30)
        self.btnDialog1.clicked.connect(self.D1key)

        self.btnDialog2 = QPushButton("Save Key", self.dialog)
        self.btnDialog2.resize(70,25)
        self.btnDialog2.move(85, 30)
        self.btnDialog2.clicked.connect(self.D2key)

        self.btnDialog3 = QPushButton("Save Key", self.dialog)
        self.btnDialog3.resize(70,25)
        self.btnDialog3.move(165, 30)
        self.btnDialog3.clicked.connect(self.D3key)

        self.btnDialog4 = QPushButton("Save Key", self.dialog)
        self.btnDialog4.resize(70,25)
        self.btnDialog4.move(245, 30)
        self.btnDialog4.clicked.connect(self.D4key)

        self.btnDialog5 = QPushButton("Save Key", self.dialog)
        self.btnDialog5.resize(70,25)
        self.btnDialog5.move(325, 30)
        self.btnDialog5.clicked.connect(self.D5key)

        self.btnDialog6 = QPushButton("Save Key", self.dialog)
        self.btnDialog6.resize(70,25)
        self.btnDialog6.move(5, 90)
        self.btnDialog6.clicked.connect(self.F1key)

        self.btnDialog6 = QPushButton("Save Key", self.dialog)
        self.btnDialog6.resize(70,25)
        self.btnDialog6.move(85, 90)
        self.btnDialog6.clicked.connect(self.F2key)

        self.btnDialog6 = QPushButton("Save Key", self.dialog)
        self.btnDialog6.resize(70,25)
        self.btnDialog6.move(165, 90)
        self.btnDialog6.clicked.connect(self.F3key)

        self.btnDialog6 = QPushButton("Save Key", self.dialog)
        self.btnDialog6.resize(70,25)
        self.btnDialog6.move(245, 90)
        self.btnDialog6.clicked.connect(self.F4key)

        self.btnDialog6 = QPushButton("Save Key", self.dialog)
        self.btnDialog6.resize(70,25)
        self.btnDialog6.move(325, 90)
        self.btnDialog6.clicked.connect(self.F5key)

        self.btnDialog7 = QPushButton("Save All", self.dialog)
        self.btnDialog7.resize(70,40)
        self.btnDialog7.move(120, 125)
        self.btnDialog7.clicked.connect(self.Save_all)

        self.btnDialog = QPushButton("Close", self.dialog)
        self.btnDialog.resize(70,40)
        self.btnDialog.move(210, 125)
        self.btnDialog.clicked.connect(self.dialog_close)

        self.dialog.setWindowTitle('Dialog')
        self.dialog.resize(400, 180)
        self.dialog.show()

    def Save_all(self):
        try:
            self.wb = openpyxl.Workbook()
            self.file = openpyxl.load_workbook("Key Data.xlsx")
            self.sheet = self.file.active
            self.key1 = self.keyy1.text()
            self.sheet['A1'] = self.key1
            self.key3 = self.keyy2.text()
            self.sheet['A2'] = self.key3
            self.key5 = self.keyy3.text()
            self.sheet['A3'] = self.key5
            self.key7 = self.keyy4.text()
            self.sheet['A4'] = self.key7
            self.key9 = self.keyy5.text()
            self.sheet['A5'] = self.key9
            self.key2 = self.keyy6.text()
            self.sheet['B1'] = self.key2
            self.key4 = self.keyy7.text()
            self.sheet['B2'] = self.key4
            self.key6 = self.keyy8.text()
            self.sheet['B3'] = self.key6
            self.key8 = self.keyy9.text()
            self.sheet['B4'] = self.key8
            self.key10 = self.keyy10.text()
            self.sheet['B5'] = self.key10
            self.wb.save("Key Data.xlsx")

            reply = QMessageBox.question(self, 'Save', 'Save Complete',
                                        QMessageBox.Yes)

            if reply == QMessageBox.Yes:
                self.dialog.close()
                pass
        except:
            pass

    def dialog_close(self):
        self.dialog.close()

    def D1key(self):
        try:
            self.key1 = self.keyy1.text()
            self.sheet['A1'] = self.key1
            self.wb.save("Key Data.xlsx")
        except:
            pass
    def D2key(self):
        try:
            self.key3 = self.keyy2.text()
            self.sheet['A2'] = self.key3
            self.wb.save("Key Data.xlsx")
        except:
            pass
    def D3key(self):
        try:
            self.key5 = self.keyy3.text()
            self.sheet['A3'] = self.key5
            self.wb.save("Key Data.xlsx")
        except:
            pass
    def D4key(self):
        try:
            self.key7 = self.keyy4.text()
            self.sheet['A4'] = self.key7
            self.wb.save("Key Data.xlsx")
        except:
            pass

    def D5key(self):
        try:
            self.key9 = self.keyy5.text()
            self.sheet['A5'] = self.key9
            self.wb.save("Key Data.xlsx")
        except:
            pass

    def F1key(self):
        try:
            self.key2 = self.keyy6.text()
            self.sheet['B1'] = self.key2
            self.wb.save("Key Data.xlsx")
        except:
            pass

    def F2key(self):
        try:
            self.key4 = self.keyy7.text()
            self.sheet['B2'] = self.key4
            self.wb.save("Key Data.xlsx")
        except:
            pass

    def F3key(self):
        try:
            self.key6 = self.keyy8.text()
            self.sheet['B3'] = self.key6
            self.wb.save("Key Data.xlsx")
        except:
            pass

    def F4key(self):
        try:
            self.key8 = self.keyy9.text()
            self.sheet['B4'] = self.key8
            self.wb.save("Key Data.xlsx")
        except:
            pass

    def F5key(self):
        try:
            self.key10 = self.keyy10.text()
            self.sheet['B5'] = self.key10
            self.wb.save("Key Data.xlsx")
        except:
            pass

    def keyEvent(self, state):
        try:
            if state == Qt.Checked:
                self.keyboard_thread = keyboard.Listener(on_press=self.keyPressEvent)
                self.keyboard_thread.start()
            else:
                self.keyboard_thread.stop()
        except:
            pass

    def ResetBtn(self):
        try:
            self.Reset.toggle()
            self.pbar.reset()
            self.pbar2.reset()
            self.pbar3.reset()
            self.pbar4.reset()
            self.pbar6.reset()
            self.pbarr.reset()
            self.pbarr2.reset()
            self.pbarr3.reset()
            self.pbarr4.reset()
            self.pbar5.reset()
            self.timer.stop()
            self.step = 0
            self.timer2.stop()
            self.step2 = 0
            self.timer3.stop()
            self.step3 = 0
            self.timer4.stop()
            self.step4 = 0
            self.timer5.stop()
            self.step5 = 0
            self.timer6.stop()
            self.step6 = 0
            self.timerr2.stop()
            self.step7 = 0
            self.timerr.stop()
            self.step8 = 0
            self.timerr3.stop()
            self.step9 = 0
            self.timerr4.stop()
            self.step10 = 0
            self.label1.setText('000')
            self.label1.repaint()
            self.label2.setText('000')
            self.label2.repaint()
            self.label3.setText('000')
            self.label3.repaint()
            self.label4.setText('000')
            self.label4.repaint()
            self.label5.setText('000')
            self.label5.repaint()
            self.label6.setText('000')
            self.label6.repaint()
            self.label7.setText('000')
            self.label7.repaint()
            self.label8.setText('000')
            self.label8.repaint()
            self.label9.setText('000')
            self.label9.repaint()
            self.label10.setText('000')
            self.label10.repaint()
            self.ttimer1.stop()
            self.ttimer2.stop()
            self.ttimer3.stop()
            self.ttimer4.stop()
            self.ttimer5.stop()
            self.ttimer6.stop()
            self.ttimer7.stop()
            self.ttimer8.stop()
            self.ttimer9.stop()
            self.ttimer10.stop()
            opacity_effect1 = QGraphicsOpacityEffect(self.Dbtn1)
            opacity_effect1.setOpacity(1)
            opacity_effect2 = QGraphicsOpacityEffect(self.Dbtn2)
            opacity_effect2.setOpacity(1)
            opacity_effect3 = QGraphicsOpacityEffect(self.Dbtn3)
            opacity_effect3.setOpacity(1)
            opacity_effect4 = QGraphicsOpacityEffect(self.Dbtn4)
            opacity_effect4.setOpacity(1)
            opacity_effect5 = QGraphicsOpacityEffect(self.Dbtn5)
            opacity_effect5.setOpacity(1)
            opacity_effect6 = QGraphicsOpacityEffect(self.Fbtn)
            opacity_effect6.setOpacity(1)
            opacity_effect7 = QGraphicsOpacityEffect(self.Fbtn2)
            opacity_effect7.setOpacity(1)
            opacity_effect8 = QGraphicsOpacityEffect(self.Fbtn3)
            opacity_effect8.setOpacity(1)
            opacity_effect9 = QGraphicsOpacityEffect(self.Fbtn4)
            opacity_effect9.setOpacity(1)
            opacity_effect10 = QGraphicsOpacityEffect(self.Fbtn5)
            opacity_effect10.setOpacity(1)
            self.Dbtn1.setGraphicsEffect(opacity_effect1)
            self.Dbtn2.setGraphicsEffect(opacity_effect2)
            self.Dbtn3.setGraphicsEffect(opacity_effect3)
            self.Dbtn4.setGraphicsEffect(opacity_effect4)
            self.Dbtn5.setGraphicsEffect(opacity_effect5)
            self.Fbtn.setGraphicsEffect(opacity_effect6)
            self.Fbtn2.setGraphicsEffect(opacity_effect7)
            self.Fbtn3.setGraphicsEffect(opacity_effect8)
            self.Fbtn4.setGraphicsEffect(opacity_effect9)
            self.Fbtn5.setGraphicsEffect(opacity_effect10)
        except:
            pass

    def GameMode(self, state):
        if state == Qt.Checked:
            self.setFixedSize(450, 113)
            self.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.FramelessWindowHint)
            self.move(5, 935)
            self.show()
        else:
            self.setWindowTitle('스펠계산기')
            self.setWindowIcon(QIcon('img\\icon.png'))
            self.setWindowFlags(Qt.WindowStaysOnTopHint)
            self.setWindowOpacity(1)
            self.setFixedSize(450, 235)
            self.resize(450, 235)
            self.center()
            self.show()

    def doac1(self, text):
        self.Dbtn1.setText(text)

    def doac2(self, text):
        self.Dbtn2.setText(text)

    def doac3(self, text):
        self.Dbtn3.setText(text)

    def doac4(self, text):
        self.Dbtn4.setText(text)

    def doac5(self, text):
        self.Dbtn5.setText(text)

    def foac1(self, text):
        self.Fbtn.setText(text)

    def foac2(self, text):
        self.Fbtn2.setText(text)

    def foac3(self, text):
        self.Fbtn3.setText(text)

    def foac4(self, text):
        self.Fbtn4.setText(text)

    def foac5(self, text):
        self.Fbtn5.setText(text)

    def setValue(self, value):
        if value == 1:
            self.setWindowOpacity(1)
        if value == 2:
            self.setWindowOpacity(0.9)
        if value == 3:
            self.setWindowOpacity(0.8)
        if value == 4:
            self.setWindowOpacity(0.7)
        if value == 5:
            self.setWindowOpacity(0.6)
        if value == 6:
            self.setWindowOpacity(0.5)
        if value == 7:
            self.setWindowOpacity(0.4)
        if value == 8:
            self.setWindowOpacity(0.3)
        if value == 9:
            self.setWindowOpacity(0.2)
        if value == 10:
            self.setWindowOpacity(0.1)

    def ddsp(self):
        try:
            for i in range(self.dcb.count()):
                if self.dcb.currentIndex() == i:
                    self.dsec = self.a[i - 1]
                    self.dsecc = self.dsec
                    self.Time1 = self.T[i - 1]
                    self.Timee1 = self.Time1
                    return
        except:
            pass

    def ddsp2(self):
        try:
            for i in range(self.dcb2.count()):
                if self.dcb2.currentIndex() == i:
                    self.dsec2 = self.a[i - 1]
                    self.dsecc2 = self.dsec2
                    self.Time2 = self.T[i - 1]
                    self.Timee2 = self.Time2
                    return
        except:
            pass

    def ddsp3(self):
        try:
            for i in range(self.dcb3.count()):
                if self.dcb3.currentIndex() == i:
                    self.dsec3 = self.a[i - 1]
                    self.dsecc3 = self.dsec3
                    self.Time3 = self.T[i - 1]
                    self.Timee3 = self.Time3
                    return
        except:
            pass

    def ddsp4(self):
        try:
            for i in range(0, self.dcb4.count()):
                if self.dcb4.currentIndex() == i:
                    self.dsec4 = self.a[i - 1]
                    self.dsecc4 = self.dsec4
                    self.Time4 = self.T[i - 1]
                    self.Timee4 = self.Time4
                    return
        except:
            pass

    def ddsp5(self):
        try:
            for i in range(0, self.dcb5.count()):
                if self.dcb5.currentIndex() == i:
                    self.dsec5 = self.a[i - 1]
                    self.dsecc5 = self.dsec5
                    self.Time5 = self.T[i - 1]
                    self.Timee5 = self.Time5
                    return
        except:
            pass

    def ffsp(self):
        try:
            for i in range(0, self.fcb.count()):
                if self.fcb.currentIndex() == i:
                    self.fsec = self.a[i - 1]
                    self.fsecc = self.fsec
                    self.Time6 = self.T[i - 1]
                    self.Timee6 = self.Time6
                    return
        except:
            pass

    def ffsp2(self):
        try:
            for i in range(0, self.fcb2.count()):
                if self.fcb2.currentIndex() == i:
                    self.fsec2 = self.a[i - 1]
                    self.fsecc2 = self.fsec2
                    self.Time7 = self.T[i - 1]
                    self.Timee7 = self.Time7
                    return
        except:
            pass

    def ffsp3(self):
        try:
            for i in range(0, self.fcb3.count()):
                if self.fcb3.currentIndex() == i:
                    self.fsec3 = self.a[i - 1]
                    self.fsecc3 = self.fsec3
                    self.Time8 = self.T[i - 1]
                    self.Timee8 = self.Time8
                    return
        except:
            pass

    def ffsp4(self):
        try:
            for i in range(0, self.fcb4.count()):
                if self.fcb4.currentIndex() == i:
                    self.fsec4 = self.a[i - 1]
                    self.fsecc4 = self.fsec4
                    self.Time9 = self.T[i - 1]
                    self.Timee9 = self.Time9
                    return
        except:
            pass

    def ffsp5(self):
        try:
            for i in range(0, self.fcb5.count()):
                if self.fcb5.currentIndex() == i:
                    self.fsec5 = self.a[i - 1]
                    self.fsecc5 = self.fsec5
                    self.Time10 = self.T[i - 1]
                    self.Timee10 = self.Time10
                    return
        except:
            pass

    def timerEvent(self):
        if self.step >= 100:
            self.timer.stop()
            self.pbar.reset()
            self.step = 0
            opacity_effect = QGraphicsOpacityEffect(self.Dbtn1)
            opacity_effect.setOpacity(1)
            self.Dbtn1.setGraphicsEffect(opacity_effect)
            return
        self.step = self.step + 1
        self.pbar.setValue(self.step)

    def timerEvent2(self):
        if self.step2 >= 100:
            self.timer2.stop()
            self.pbar2.reset()
            self.step2 = 0
            opacity_effect = QGraphicsOpacityEffect(self.Fbtn)
            opacity_effect.setOpacity(1)
            self.Fbtn.setGraphicsEffect(opacity_effect)
            return
        self.step2 = self.step2 + 1
        self.pbar2.setValue(self.step2)

    def timerEvent3(self):
        if self.step3 >= 100:
            self.timer3.stop()
            self.pbar3.reset()
            self.step3 = 0
            opacity_effect = QGraphicsOpacityEffect(self.Dbtn2)
            opacity_effect.setOpacity(1)
            self.Dbtn2.setGraphicsEffect(opacity_effect)
            return
        self.step3 = self.step3 + 1
        self.pbar3.setValue(self.step3)

    def timerEvent4(self):
        if self.step4 >= 100:
            self.timer4.stop()
            self.pbar4.reset()
            self.step4 = 0
            opacity_effect = QGraphicsOpacityEffect(self.Dbtn3)
            opacity_effect.setOpacity(1)
            self.Dbtn3.setGraphicsEffect(opacity_effect)
            return
        self.step4 = self.step4 + 1
        self.pbar4.setValue(self.step4)

    def timerEvent5(self):
        if self.step5 >= 100:
            self.timer5.stop()
            self.pbar5.reset()
            self.step5 = 0
            opacity_effect = QGraphicsOpacityEffect(self.Dbtn4)
            opacity_effect.setOpacity(1)
            self.Dbtn4.setGraphicsEffect(opacity_effect)
            return
        self.step5 = self.step5 + 1
        self.pbar5.setValue(self.step5)

    def timerEvent6(self):
        if self.step6 >= 100:
            self.timer6.stop()
            self.pbar6.reset()
            self.step6 = 0
            opacity_effect = QGraphicsOpacityEffect(self.Dbtn5)
            opacity_effect.setOpacity(1)
            self.Dbtn5.setGraphicsEffect(opacity_effect)
            return
        self.step6 = self.step6 + 1
        self.pbar6.setValue(self.step6)

    def timerEvent7(self):
        if self.stepp >= 100:
            self.timerr.stop()
            self.pbarr.reset()
            self.stepp = 0
            opacity_effect = QGraphicsOpacityEffect(self.Fbtn2)
            opacity_effect.setOpacity(1)
            self.Fbtn2.setGraphicsEffect(opacity_effect)
            return
        self.stepp = self.stepp + 1
        self.pbarr.setValue(self.stepp)

    def timerEvent8(self):
        if self.stepp2 >= 100:
            self.timerr2.stop()
            self.pbarr2.reset()
            self.stepp2 = 0
            opacity_effect = QGraphicsOpacityEffect(self.Fbtn3)
            opacity_effect.setOpacity(1)
            self.Fbtn3.setGraphicsEffect(opacity_effect)
            return
        self.stepp2 = self.stepp2 + 1
        self.pbarr2.setValue(self.stepp2)

    def timerEvent9(self):
        if self.stepp3 >= 100:
            self.timerr3.stop()
            self.pbarr3.reset()
            self.stepp3 = 0
            opacity_effect = QGraphicsOpacityEffect(self.Fbtn4)
            opacity_effect.setOpacity(1)
            self.Fbtn4.setGraphicsEffect(opacity_effect)
            return
        self.stepp3 = self.stepp3 + 1
        self.pbarr3.setValue(self.stepp3)

    def timerEvent10(self):
        if self.stepp4 >= 100:
            self.timerr4.stop()
            self.pbarr4.reset()
            self.stepp4 = 0
            opacity_effect = QGraphicsOpacityEffect(self.Fbtn5)
            opacity_effect.setOpacity(1)
            self.Fbtn5.setGraphicsEffect(opacity_effect)
            return
        self.stepp4 = self.stepp4 + 1
        self.pbarr4.setValue(self.stepp4)

    def time1(self):
        try:
            if self.val <= 0:
                self.ttimer1.stop()
                return
            self.val = self.val - 1
            self.label1.setText(str(self.val))
            self.label1.repaint()
        except:
            pass

    def timer1(self):
        self.ttimer1 = QTimer()
        self.ttimer1.start(1000)
        self.ttimer1.timeout.connect(self.time1)

    def time2(self):
        try:
            if self.val2 <= 0:
                self.ttimer2.stop()
                return
            self.val2 = self.val2 - 1
            self.label2.setText(str(self.val2))
            self.label2.repaint()
        except:
            pass

    def timer2(self):
        self.ttimer2 = QTimer()
        self.ttimer2.start(1000)
        self.ttimer2.timeout.connect(self.time2)

    def time3(self):
        try:
            if self.val3 <= 0:
                self.ttimer3.stop()
                return
            self.val3 = self.val3 - 1
            self.label3.setText(str(self.val3))
            self.label3.repaint()
        except:
            pass

    def timer3(self):
        self.ttimer3 = QTimer()
        self.ttimer3.start(1000)
        self.ttimer3.timeout.connect(self.time3)

    def time4(self):
        try:
            if self.val4 <= 0:
                self.ttimer4.stop()
                return
            self.val4 = self.val4 - 1
            self.label4.setText(str(self.val4))
            self.label4.repaint()
        except:
            pass

    def timer4(self):
        self.ttimer4 = QTimer()
        self.ttimer4.start(1000)
        self.ttimer4.timeout.connect(self.time4)

    def time5(self):
        try:
            if self.val5 <= 0:
                self.ttimer5.stop()
                return
            self.val5 = self.val5 - 1
            self.label5.setText(str(self.val5))
            self.label5.repaint()
        except:
            pass

    def timer5(self):
        self.ttimer5 = QTimer()
        self.ttimer5.start(1000)
        self.ttimer5.timeout.connect(self.time5)

    def time6(self):
        try:
            if self.val6 <= 0:
                self.ttimer6.stop()
                return
            self.val6 = self.val6 - 1
            self.label6.setText(str(self.val6))
            self.label6.repaint()
        except:
            pass

    def timer6(self):
        self.ttimer6 = QTimer()
        self.ttimer6.start(1000)
        self.ttimer6.timeout.connect(self.time6)

    def time7(self):
        try:
            if self.val7 <= 0:
                self.ttimer7.stop()
                return
            self.val7 = self.val7 - 1
            self.label7.setText(str(self.val7))
            self.label7.repaint()
        except:
            pass

    def timer7(self):
        self.ttimer7 = QTimer()
        self.ttimer7.start(1000)
        self.ttimer7.timeout.connect(self.time7)

    def time8(self):
        try:
            if self.val8 <= 0:
                self.ttimer8.stop()
                return
            self.val8 = self.val8 - 1
            self.label8.setText(str(self.val8))
            self.label8.repaint()
        except:
            pass

    def timer8(self):
        self.ttimer8 = QTimer()
        self.ttimer8.start(1000)
        self.ttimer8.timeout.connect(self.time8)

    def time9(self):
        try:
            if self.val9 <= 0:
                self.ttimer9.stop()
                return
            self.val9 = self.val9 - 1
            self.label9.setText(str(self.val9))
            self.label9.repaint()
        except:
            pass

    def timer9(self):
        self.ttimer9 = QTimer()
        self.ttimer9.start(1000)
        self.ttimer9.timeout.connect(self.time9)

    def time10(self):
        try:
            if self.val10 <= 0:
                self.ttimer10.stop()
                return
            self.val10 = self.val10 - 1
            self.label10.setText(str(self.val10))
            self.label10.repaint()
        except:
            pass

    def timer10(self):
        self.ttimer10 = QTimer()
        self.ttimer10.start(1000)
        self.ttimer10.timeout.connect(self.time10)

    def Dbtnn(self):
        try:
            self.Dbtn1.toggle()
            self.timer = QTimer()
            self.step = 0
            self.timer.timeout.connect(self.timerEvent)
            self.timer.start(int(self.dsecc))
            opacity_effect = QGraphicsOpacityEffect(self.Dbtn1)
            opacity_effect.setOpacity(0.0)
            self.Dbtn1.setGraphicsEffect(opacity_effect)
            self.val = (int(self.Timee1))
        except:
            pass

    def Dbtnn2(self):
        try:
            self.Dbtn2.toggle()
            self.timer3 = QTimer()
            self.step3 = 0
            self.timer3.timeout.connect(self.timerEvent3)
            self.timer3.start(int(self.dsecc2))
            opacity_effect = QGraphicsOpacityEffect(self.Dbtn2)
            opacity_effect.setOpacity(0.0)
            self.Dbtn2.setGraphicsEffect(opacity_effect)
            self.val2 = (int(self.Timee2))
        except:
            pass

    def Dbtnn3(self):
        try:
            self.Dbtn3.toggle()
            self.timer4 = QTimer()
            self.step4 = 0
            self.timer4.timeout.connect(self.timerEvent4)
            self.timer4.start(int(self.dsecc3))
            opacity_effect = QGraphicsOpacityEffect(self.Dbtn3)
            opacity_effect.setOpacity(0.0)
            self.Dbtn3.setGraphicsEffect(opacity_effect)
            self.val3 = (int(self.Timee3))
        except:
            pass

    def Dbtnn4(self):
        try:
            self.Dbtn4.toggle()
            self.timer5 = QTimer()
            self.step5 = 0
            self.timer5.timeout.connect(self.timerEvent5)
            self.timer5.start(int(self.dsecc4))
            opacity_effect = QGraphicsOpacityEffect(self.Dbtn4)
            opacity_effect.setOpacity(0.0)
            self.Dbtn4.setGraphicsEffect(opacity_effect)
            self.val4 = (int(self.Timee4))
        except:
            pass

    def Dbtnn5(self):
        try:
            self.Dbtn5.toggle()
            self.timer6 = QTimer()
            self.step6 = 0
            self.timer6.timeout.connect(self.timerEvent6)
            self.timer6.start(int(self.dsecc5))
            opacity_effect = QGraphicsOpacityEffect(self.Dbtn5)
            opacity_effect.setOpacity(0.0)
            self.Dbtn5.setGraphicsEffect(opacity_effect)
            self.val5 = (int(self.Timee5))
        except:
            pass

    def Fbtnn(self):
        try:
            self.Fbtn.toggle()
            self.timer2 = QTimer()
            self.step2 = 0
            self.timer2.start(int(self.fsecc))
            opacity_effect = QGraphicsOpacityEffect(self.Fbtn)
            opacity_effect.setOpacity(0.0)
            self.Fbtn.setGraphicsEffect(opacity_effect)
            self.timer2.timeout.connect(self.timerEvent2)
            self.val6 = (int(self.Timee6))
        except:
            pass

    def Fbtnn2(self):
        try:
            self.Fbtn2.toggle()
            self.timerr = QTimer()
            self.stepp = 0
            self.timerr.timeout.connect(self.timerEvent7)
            self.timerr.start(int(self.fsecc2))
            opacity_effect = QGraphicsOpacityEffect(self.Fbtn2)
            opacity_effect.setOpacity(0.0)
            self.Fbtn2.setGraphicsEffect(opacity_effect)
            self.val7 = (int(self.Timee7))
        except:
            pass

    def Fbtnn3(self):
        try:
            self.Fbtn3.toggle()
            self.timerr2 = QTimer()
            self.stepp2 = 0
            self.timerr2.timeout.connect(self.timerEvent8)
            self.timerr2.start(int(self.fsecc3))
            opacity_effect = QGraphicsOpacityEffect(self.Fbtn3)
            opacity_effect.setOpacity(0.0)
            self.Fbtn3.setGraphicsEffect(opacity_effect)
            self.val8 = (int(self.Timee8))
        except:
            pass

    def Fbtnn4(self):
        try:
            self.Fbtn4.toggle()
            self.timerr3 = QTimer()
            self.stepp3 = 0
            self.timerr3.timeout.connect(self.timerEvent9)
            self.timerr3.start(int(self.fsecc4))
            opacity_effect = QGraphicsOpacityEffect(self.Fbtn4)
            opacity_effect.setOpacity(0.0)
            self.Fbtn4.setGraphicsEffect(opacity_effect)
            self.val9 = (int(self.Timee9))
        except:
            pass

    def Fbtnn5(self):
        try:
            self.Fbtn5.toggle()
            self.timerr4 = QTimer()
            self.stepp4 = 0
            self.timerr4.timeout.connect(self.timerEvent10)
            self.timerr4.start(int(self.fsecc5))
            opacity_effect = QGraphicsOpacityEffect(self.Fbtn5)
            opacity_effect.setOpacity(0.0)
            self.Fbtn5.setGraphicsEffect(opacity_effect)
            self.val10 = (int(self.Timee10))
        except:
            pass

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def keyPressEvent(self, Key):
        while True:
            try:
                if self.key1 == Key.char:
                    self.Dbtn1.click()
                    return

                if self.key2 == Key.char:
                    self.Fbtn.click()
                    return

                if self.key3 == Key.char:
                    self.Dbtn2.click()
                    return

                if self.key4 == Key.char:
                    self.Fbtn2.click()
                    return

                if self.key5 == Key.char:
                    self.Dbtn3.click()
                    return

                if self.key6 == Key.char:
                    self.Fbtn3.click()
                    return

                if self.key7 == Key.char:
                    self.Dbtn4.click()
                    return

                if self.key8 == Key.char:
                    self.Fbtn4.click()
                    return

                if self.key9 == Key.char:
                    self.Dbtn5.click()
                    return

                if self.key10 == Key.char:
                    self.Fbtn5.click()
                    return
                return
            except:
                return

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())
